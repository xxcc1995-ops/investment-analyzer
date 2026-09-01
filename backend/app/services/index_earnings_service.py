# -*- coding: utf-8 -*-
"""指数盈利与估值服务

读取用户手工维护的三大指数「盈利与估值分析表」Excel（周度数据），
解析为标准 JSON 供前端表格 + 折线图（划线）查阅。

数据特征（2026-08-15 拆解确认）：
- 每个文件两个 sheet：Sheet1 = 说明行 + 表头行 + 周度数据 + 表尾(EPS周期表/统计摘要)；
  「曲线图」sheet 通常为嵌入图表（无单元格数据），万得全A 的 EPS 周期表在该 sheet。
- Sheet1 第1行：数据来源与口径说明（长文本）；第2行：表头；第3行起：周度数据。
- 口径：估值中枢偏移率 = PE-TTM × 十年期国债收益率（股债收益比倒数×100），
  基准线 标普500=70(7折) / 万得全A=60(6折) / 沪深300=50(5折)；
  合理收盘价 = (100÷十年期国债) × 折扣 × 隐含EPS(收盘价÷PE)；
  风险溢价 = 100÷PE − 十年期国债收益率（百分点）。
- EPS 周期：隐含EPS 平滑 + 4% zigzag 阈值划分上涨(红)/下降(绿)周期。

缓存策略：按文件 mtime 失效——用户更新 Excel 后下次请求自动重读。
"""
import logging
import os
from datetime import date, datetime

import openpyxl

logger = logging.getLogger(__name__)

# ============ 指数配置 ============

INDEX_EARNINGS_CONFIG = {
    "sp500": {
        "name": "标普500",
        "file": r"D:/1957~2026年标普500盈利与估值/标普500_1957-2026盈利与估值分析表.xlsx",
        "market": "美股",
        "baseline": 70,          # 估值中枢偏移率基准线（7折）
        "discount": 0.7,
        "bond_name": "美国十年期国债收益率",
    },
    "wind_all_a": {
        "name": "万得全A",
        "file": r"D:/1999年~2025年万得全A盈利与估值/万得全A1999~2026盈利与估值分析表.xlsx",
        "market": "A股",
        "baseline": 60,          # 6折
        "discount": 0.6,
        "bond_name": "十年期中国国债收益率",
    },
    "hs300": {
        "name": "沪深300",
        "file": r"D:/2005年~2026年沪深300盈利与估值/沪深300_2005-2026盈利与估值分析表.xlsx",
        "market": "A股",
        "baseline": 50,          # 5折
        "discount": 0.5,
        "bond_name": "十年期中国国债收益率",
    },
}

# 字段标准名 → 表格展示标签（顺序即表格列顺序）
FIELD_LABELS = [
    ("date", "交易日期"),
    ("close", "收盘价"),
    ("pe", "PE-TTM"),
    ("risk_premium", "风险溢价(百分点)"),
    ("eps_ttm", "EPS(TTM)"),
    ("implied_eps", "隐含EPS"),
    ("eps_up", "EPS上涨周期(红)"),
    ("eps_down", "EPS下降周期(绿)"),
    ("us_cn_spread", "中美国债利差(×20)"),
    ("cn10y", "中国十年期国债(%)"),
    ("us10y", "美国十年期国债(%)"),
    ("valuation_dev", "估值中枢偏移率"),
    ("fair_close", "合理收盘价"),
    ("pb", "市净率PB"),
    ("up_time", "盈利上涨总时间"),
    ("down_time", "盈利下降总时间"),
]

# 数据表中默认隐藏的过程列（仅图表/统计用，减少表格噪音）
TABLE_HIDDEN_FIELDS = {"eps_up", "eps_down", "up_time", "down_time"}


def _map_header(raw) -> str | None:
    """把 Excel 表头（含换行符）映射为标准字段名。"""
    if raw is None:
        return None
    t = str(raw).replace("\n", "").replace(" ", "").replace("　", "")
    if t == "交易日期":
        return "date"
    if t == "收盘价":
        return "close"
    if "市盈率" in t:
        return "pe"
    if "风险溢价" in t:
        return "risk_premium"
    if t.startswith("EPS") and "TTM" in t:
        return "eps_ttm"
    if "增长周期" in t:
        return "eps_up"
    if "下降周期" in t:
        return "eps_down"
    if "中美国债利差" in t:
        return "us_cn_spread"
    if "十年期中国" in t:
        return "cn10y"
    if "美国十年期" in t:
        return "us10y"
    if "估值中枢" in t:
        return "valuation_dev"
    if "盈利上涨" in t:
        return "up_time"
    if "盈利下降" in t:
        return "down_time"
    if "合理收盘价" in t:
        return "fair_close"
    if "市净率" in t:
        return "pb"
    if "隐含EPS" in t:
        return "implied_eps"
    return None


def _to_date_str(v) -> str | None:
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, date):
        return v.isoformat()
    return None


def _num(v):
    """数值清洗：保留 None，数值四舍五入到4位小数。"""
    if v is None or isinstance(v, str):
        return None
    if isinstance(v, (int, float)):
        return round(v, 4)
    return None


def _is_cycle_row(row) -> bool:
    """EPS 周期表行特征：第3列为 上涨/下降。"""
    return (
        row is not None
        and len(row) >= 3
        and isinstance(row[2], str)
        and row[2].strip() in ("上涨", "下降")
    )


def _parse_cycle_row(row) -> dict:
    def _s(v):
        if isinstance(v, datetime):
            return v.strftime("%Y-%m-%d")
        return str(v).strip() if v is not None else None

    return {
        "start": _s(row[0]),
        "end": _s(row[1]),
        "direction": row[2].strip(),
        "months": _s(row[3]) if len(row) > 3 else None,
        "weeks": _s(row[4]) if len(row) > 4 else None,
    }


_CACHE: dict[str, tuple[float, dict]] = {}  # code -> (mtime, payload)


def _load_index(code: str) -> dict:
    """加载并解析单个指数 Excel，按文件 mtime 缓存。"""
    cfg = INDEX_EARNINGS_CONFIG[code]
    path = cfg["file"]
    if not os.path.exists(path):
        raise FileNotFoundError(f"数据文件不存在: {path}")

    mtime = os.path.getmtime(path)
    cached = _CACHE.get(code)
    if cached and cached[0] == mtime:
        return cached[1]

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))

        # 第1行：数据来源与口径说明（第2列是长文本）
        notes = ""
        if rows and len(rows[0]) > 1 and isinstance(rows[0][1], str):
            notes = rows[0][1].strip()

        # 第2行：表头
        header = rows[1] if len(rows) > 1 else ()
        col_fields = [_map_header(h) for h in header]

        data_rows: list[dict] = []
        cycles: list[dict] = []
        summary_lines: list[str] = []

        for row in rows[2:]:
            if not row or row[0] is None:
                continue
            d = _to_date_str(row[0])
            if d:
                rec = {}
                for idx, field in enumerate(col_fields):
                    if field is None or field == "date" or idx >= len(row):
                        continue
                    rec[field] = _num(row[idx])
                rec["date"] = d
                data_rows.append(rec)
            elif _is_cycle_row(row):
                cycles.append(_parse_cycle_row(row))
            elif isinstance(row[0], str):
                text = row[0].strip()
                if text:
                    summary_lines.append(text)

        # EPS 周期表兜底：Sheet1 没找到时去「曲线图」sheet 找（万得全A 在该 sheet）
        if not cycles and len(wb.sheetnames) > 1:
            ws2 = wb[wb.sheetnames[1]]
            for row in ws2.iter_rows(values_only=True):
                if _is_cycle_row(row):
                    cycles.append(_parse_cycle_row(row))

        # 可用字段（按展示顺序）
        present = {k for r in data_rows for k in r.keys()}
        fields = [f for f, _ in FIELD_LABELS if f in present]
        columns = [
            {"key": f, "label": lbl}
            for f, lbl in FIELD_LABELS
            if f in present and f not in TABLE_HIDDEN_FIELDS
        ]

        # 盈利上涨/下降总时间（只在首行出现的汇总统计）→ 提到 meta
        up_total = next((r["up_time"] for r in data_rows if r.get("up_time") is not None), None)
        down_total = next((r["down_time"] for r in data_rows if r.get("down_time") is not None), None)

        latest = data_rows[-1] if data_rows else {}
        payload = {
            "meta": {
                "code": code,
                "name": cfg["name"],
                "market": cfg["market"],
                "baseline": cfg["baseline"],
                "discount": cfg["discount"],
                "bond_name": cfg["bond_name"],
                "start_date": data_rows[0]["date"] if data_rows else None,
                "end_date": latest.get("date"),
                "row_count": len(data_rows),
                "file": path,
                "file_updated": datetime.fromtimestamp(mtime).strftime("%Y-%m-%d %H:%M:%S"),
                "up_total_time": up_total,
                "down_total_time": down_total,
                "latest": {
                    "date": latest.get("date"),
                    "close": latest.get("close"),
                    "pe": latest.get("pe"),
                    "risk_premium": latest.get("risk_premium"),
                    "valuation_dev": latest.get("valuation_dev"),
                    "fair_close": latest.get("fair_close"),
                },
            },
            "notes": notes,
            "summary_lines": summary_lines,
            "fields": fields,
            "columns": columns,
            "rows": data_rows,
            "cycles": cycles,
        }
        _CACHE[code] = (mtime, payload)
        logger.info(
            f"[index-earnings] 加载 {cfg['name']}: {len(data_rows)} 行, "
            f"{len(cycles)} 个EPS周期, 区间 {payload['meta']['start_date']} ~ {payload['meta']['end_date']}"
        )
        return payload
    finally:
        wb.close()


def get_index_list() -> list[dict]:
    """指数列表（含元信息，不含数据行）。文件缺失时标记 error 而不是抛异常。"""
    result = []
    for code in INDEX_EARNINGS_CONFIG:
        try:
            payload = _load_index(code)
            result.append(payload["meta"])
        except Exception as e:
            cfg = INDEX_EARNINGS_CONFIG[code]
            logger.warning(f"[index-earnings] {cfg['name']} 加载失败: {e}")
            result.append({"code": code, "name": cfg["name"], "market": cfg["market"], "error": str(e)})
    return result


def get_index_data(code: str) -> dict | None:
    """单个指数完整数据（表格 + 图表 + EPS周期 + 口径说明）。"""
    if code not in INDEX_EARNINGS_CONFIG:
        return None
    return _load_index(code)
