# -*- coding: utf-8 -*-
"""一次性分析脚本：拆解三个指数盈利与估值 Excel 的结构（只读，不修改文件）"""
import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

import openpyxl

FILES = [
    r"D:/1957~2026年标普500盈利与估值/标普500_1957-2026盈利与估值分析表.xlsx",
    r"D:/1999年~2025年万得全A盈利与估值/万得全A1999~2026盈利与估值分析表.xlsx",
    r"D:/2005年~2026年沪深300盈利与估值/沪深300_2005-2026盈利与估值分析表.xlsx",
]

for path in FILES:
    print("=" * 90)
    print("文件:", path)
    wb_f = openpyxl.load_workbook(path, data_only=False)
    wb_v = openpyxl.load_workbook(path, data_only=True)

    for name in wb_f.sheetnames:
        ws_f, ws_v = wb_f[name], wb_v[name]
        rows_f = list(ws_f.iter_rows(values_only=True))
        rows_v = list(ws_v.iter_rows(values_only=True))
        n_rows = len(rows_f)
        n_cols = max((len(r) for r in rows_f if r), default=0)
        print(f"\n--- Sheet: [{name}] | 行数: {n_rows} | 列数: {n_cols}")

        header = rows_f[0] if rows_f else []
        print("列头:", header)

        # 公式列检测（扫前10个数据行）
        formula_info = {}
        for r in rows_f[1:11]:
            if not r:
                continue
            for c in range(min(len(r), n_cols)):
                cell = r[c]
                if isinstance(cell, str) and cell.startswith("="):
                    key = str(header[c]) if c < len(header) else f"col{c}"
                    if key not in formula_info:
                        formula_info[key] = cell
        if formula_info:
            print("公式列(字段 -> 公式样例):")
            for k, v in formula_info.items():
                print(f"   {k} -> {v}")
        else:
            print("公式列: 无（均为静态值）")

        print("前5行(数值):")
        for r in rows_v[:5]:
            print("  ", r)
        print("后5行(数值):")
        for r in rows_v[-5:]:
            print("  ", r)

        # 缺失值统计
        if n_rows > 1:
            miss = [0] * n_cols
            for r in rows_v[1:]:
                if not r:
                    continue
                for c in range(min(len(r), n_cols)):
                    if r[c] is None:
                        miss[c] += 1
            h = [str(x) for x in header[:n_cols]]
            print("各列缺失数:", dict(zip(h, miss)))
    wb_f.close()
    wb_v.close()
